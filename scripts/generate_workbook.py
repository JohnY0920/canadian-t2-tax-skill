#!/usr/bin/env python3
"""
Generate a T2 Tax Preparation Excel Workbook.

This script creates a comprehensive Excel workbook from extracted financial data.
It is designed to be PARAMETERIZED -- the calling model fills in the data dictionary
before running.

Usage:
    1. Fill in the DATA dictionary below with extracted values
    2. Run: python3 generate_workbook.py

Output: [CompanyName]_Tax_[Years].xlsx in the specified output directory
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════
# FILL IN THIS DATA DICTIONARY BEFORE RUNNING
# The model should populate all fields from extracted PDF data
# ══════════════════════════════════════════════════════════════════════

DATA = {
    "company_name": "REPLACE_COMPANY_NAME",
    "business_number": "REPLACE_BN",
    "fiscal_years": [2024],  # List of years to include
    "output_dir": "/sessions/confident-kind-keller/mnt/REPLACE_FOLDER/",

    # ── Opening Position (from prior year financial statements) ──
    "opening": {
        "year": 2023,
        "balance_sheet": {
            "cash": 0,
            "prepaid_expenses": 0,
            "hst_receivable": 0,
            "computer_equipment_net": 0,  # Net book value (UCC)
            "computer_equipment_gross": 0,
            "accumulated_amortization": 0,
            "accounts_payable": 0,
            "credit_card_payable": 0,
            "hst_payable": 0,
            "due_to_shareholders": 0,
            "common_shares": 100,
            "retained_earnings": 0,
        },
        "income_statement": {
            "revenue": 0,
            "total_expenses": 0,
            "net_income": 0,  # revenue - total_expenses
        },
        "prior_year_comparatives": {
            # Optional: year before opening (e.g., 2022 if opening is 2023)
            "year": 2022,
            "revenue": 0,
            "total_expenses": 0,
            "net_income": 0,
        }
    },

    # ── Per Fiscal Year Data ──
    # Add one entry per fiscal year being filed
    "years": {
        2024: {
            # Bank statement data
            "bank_opening_balance": 0,
            "bank_closing_balance": 0,
            "bank_account_number": "",
            "bank_branch": "",

            # Revenue deposits
            "revenue_deposits": [
                # {"date": "2024-01-15", "amount": 750, "description": "Invoice - Client Name", "invoiced": True, "hst_explicit": 86.28},
                # {"date": "2024-03-05", "amount": 50, "description": "Consulting income", "invoiced": False, "hst_explicit": None},
            ],

            # Expense categories (amounts INCLUDING HST unless noted)
            "expenses": {
                "software_subscriptions": {"total": 0, "t2_line": "L8810", "deductible_pct": 1.0},
                "professional_development": {"total": 0, "t2_line": "L8670", "deductible_pct": 1.0},
                "business_registration": {"total": 0, "t2_line": "L8860", "deductible_pct": 1.0},
                "transportation_tolls": {"total": 0, "t2_line": "L9220", "deductible_pct": 1.0},
                "transportation_transit": {"total": 0, "t2_line": "L9220", "deductible_pct": 1.0},
                "transportation_parking": {"total": 0, "t2_line": "L9220", "deductible_pct": 1.0},
                "bank_charges": {"total": 0, "t2_line": "L8710", "deductible_pct": 1.0},
                "meals_entertainment": {"total": 0, "t2_line": "L8523", "deductible_pct": 0.5},
                "motor_vehicle_lease": {"total": 0, "t2_line": "L9281", "deductible_pct": 1.0},
                "motor_vehicle_repairs": {"total": 0, "t2_line": "L9225", "deductible_pct": 1.0},
                "professional_fees": {"total": 0, "t2_line": "L8860", "deductible_pct": 1.0},
                "travel": {"total": 0, "t2_line": "L9200", "deductible_pct": 1.0},
                "telephone": {"total": 0, "t2_line": "L8811", "deductible_pct": 1.0},
                "rent": {"total": 0, "t2_line": "L8910", "deductible_pct": 1.0},
                # Add more categories as needed
            },

            # Capital assets (CCA Schedule 8)
            "capital_assets": [
                # {"name": "MacBook", "class": 50, "rate": 0.55, "cost": 4756.17, "business_pct": 1.0},
                # {"name": "3D Printer", "class": 8, "rate": 0.20, "cost": 941.36, "business_pct": 0.75},
            ],

            # CCA opening UCC by class (from prior year)
            "opening_ucc": {
                # 50: 377,  # Class 50 opening UCC
                # 8: 0,     # Class 8 opening UCC
            },

            # Bank transactions (full list for bank statement sheet)
            "bank_transactions": [
                # {"month": "January", "date": "01/10/2024", "description": "Tesla Lease", "withdrawal": 865.93, "deposit": None, "category": "Motor Vehicle", "treatment": "100% deductible"},
            ],

            # Credit card monthly breakdown
            "cc_monthly": [
                # {"month": "January", "charges": 173.37, "credits": -442.10, "net": -268.73},
            ],

            # Shareholder transfers (to personal accounts)
            "shareholder_transfers": [
                # {"date": "2024-07-02", "amount": 100, "description": "Transfer to personal"},
            ],

            # Non-deductible / personal expenses
            "personal_expenses": 0,

            # Non-capital loss from prior years
            "prior_ncl": 0,  # Available NCL before this year
        },
    },

    # ── Action Items & Notes ──
    "action_items": [
        # {"item": "T2 2024 -- URGENT", "status": "File ASAP", "details": "..."},
    ],

    # ── Clarification Questions (generated by Phase 1) ──
    "clarification_questions": [
        # {"question": "What is the source of $3,420.34 deposit on Feb 10?", "category": "Unidentified Deposit"},
    ],
}


# ══════════════════════════════════════════════════════════════════════
# WORKBOOK GENERATION CODE (do not modify unless adapting)
# ══════════════════════════════════════════════════════════════════════

# Style definitions
NAVY = Font(name='Arial', bold=True, color='1F4E79')
NAVY_SMALL = Font(name='Arial', size=10, color='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
BODY_FONT = Font(name='Arial', size=11)
BLUE_INPUT = Font(name='Arial', size=11, color='0000FF')  # User-editable inputs
FLAG_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
SUBTOTAL_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
CURRENCY_FMT = '$#,##0.00'
CURRENCY_NEG = '$#,##0.00;($#,##0.00);"-"'
PCT_FMT = '0%'


def apply_header(ws, row, cols):
    """Apply header styling to a row."""
    for col_idx, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def write_row(ws, row, values, font=BODY_FONT, fill=None, number_format=None):
    """Write a data row with styling."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if number_format and isinstance(val, (int, float)):
            cell.number_format = number_format
        if isinstance(val, (int, float)) and col_idx > 1:
            cell.alignment = Alignment(horizontal='right')


def auto_width(ws, min_width=12, max_width=50):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def build_opening_position(wb, data):
    """Sheet 1: Opening Position from prior year."""
    ws = wb.create_sheet("Opening Position")
    d = data["opening"]
    bs = d["balance_sheet"]
    inc = d["income_statement"]

    r = 1
    ws.cell(row=r, column=1, value=f"{data['company_name']} -- {d['year']} Year-End (Opening Position)").font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    r += 2

    headers = ["Item", f"{d['year']} ($)", "Notes"]
    apply_header(ws, r, headers)
    r += 1

    # Balance sheet
    ws.cell(row=r, column=1, value="BALANCE SHEET").font = Font(name='Arial', bold=True, size=11, color='1F4E79')
    r += 1

    bs_items = [
        ("Cash", bs["cash"], "Bank balance at year end"),
        ("Prepaid Expenses", bs["prepaid_expenses"], ""),
        ("HST Receivable", bs["hst_receivable"], ""),
        ("Computer Equipment (Net)", bs["computer_equipment_net"], "Opening UCC for CCA"),
        ("Accounts Payable", bs["accounts_payable"], ""),
        ("Credit Card Payable", bs["credit_card_payable"], ""),
        ("HST Payable", bs["hst_payable"], ""),
        ("Due to Shareholders", bs["due_to_shareholders"], "Shareholder loan balance"),
        ("Common Shares", bs["common_shares"], ""),
        ("Retained Earnings", bs["retained_earnings"], "Opening RE for next year"),
    ]

    for item, amount, note in bs_items:
        write_row(ws, r, [item, amount, note], font=BLUE_INPUT, number_format=CURRENCY_FMT)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="INCOME STATEMENT").font = Font(name='Arial', bold=True, size=11, color='1F4E79')
    r += 1

    inc_items = [
        ("Revenue", inc["revenue"], ""),
        ("Total Expenses", inc["total_expenses"], ""),
        ("Net Income / (Loss)", inc["net_income"], ""),
    ]
    for item, amount, note in inc_items:
        write_row(ws, r, [item, amount, note], font=BLUE_INPUT, number_format=CURRENCY_NEG)
        r += 1

    auto_width(ws)
    ws.freeze_panes = 'A4'


def build_income_summary(wb, data):
    """Sheet 2: Income & Invoice Summary."""
    ws = wb.create_sheet("Income & Invoice Summary")

    r = 1
    ws.cell(row=r, column=1, value=f"{data['company_name']} -- Income Reconciliation").font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    r += 2

    headers = ["Date", "Gross ($)", "HST ($)", "Net ($)", "Source/Invoice", "Invoiced?", "Notes"]
    apply_header(ws, r, headers)
    r += 1

    for year, yd in data["years"].items():
        ws.cell(row=r, column=1, value=f"-- {year} --").font = Font(name='Arial', bold=True, size=11)
        r += 1

        total_gross = 0
        for dep in yd["revenue_deposits"]:
            gross = dep["amount"]
            if dep.get("hst_explicit") is not None:
                hst = dep["hst_explicit"]
                net = gross - hst
            else:
                net = round(gross / 1.13, 2)
                hst = round(gross - net, 2)

            invoiced = "Yes" if dep.get("invoiced") else "No -- prepare invoice"
            write_row(ws, r, [dep["date"], gross, hst, net, dep["description"], invoiced, dep.get("notes", "")],
                      number_format=CURRENCY_FMT)
            if not dep.get("invoiced"):
                ws.cell(row=r, column=6).fill = FLAG_FILL
            total_gross += gross
            r += 1

        write_row(ws, r, [f"TOTAL {year}", total_gross, "", "", "", "", ""],
                  font=Font(name='Arial', bold=True, size=11), fill=TOTAL_FILL, number_format=CURRENCY_FMT)
        r += 2

    auto_width(ws)
    ws.freeze_panes = 'A4'


def build_tax_summary(wb, data, year):
    """Sheet: Tax Summary for a specific year."""
    ws = wb.create_sheet(f"{year} Summary")
    yd = data["years"][year]

    r = 1
    ws.cell(row=r, column=1, value=f"{data['company_name']} -- {year} Tax Expense Summary (T2)").font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    r += 2

    # Expense summary
    headers = ["Category", "Tax Treatment", "T2 Line", "Total Spent ($)", "Deductible ($)", "Notes"]
    apply_header(ws, r, headers)
    r += 1

    total_spent = 0
    total_deductible = 0
    for cat_name, cat_data in yd["expenses"].items():
        if cat_data["total"] == 0:
            continue
        nice_name = cat_name.replace("_", " ").title()
        pct = cat_data["deductible_pct"]
        pct_label = f"{int(pct * 100)}%" if pct < 1 else "100%"
        deductible = round(cat_data["total"] * pct, 2)
        write_row(ws, r, [nice_name, pct_label, cat_data["t2_line"], cat_data["total"], deductible, ""],
                  font=BLUE_INPUT, number_format=CURRENCY_FMT)
        total_spent += cat_data["total"]
        total_deductible += deductible
        r += 1

    # Capital assets (not in expense total -- CCA instead)
    for asset in yd["capital_assets"]:
        write_row(ws, r, [f"Capital: {asset['name']} (CCA Class {asset['class']})", "CCA", "Sched 8",
                          asset["cost"], 0, f"Claim via Schedule 8. {int(asset['business_pct']*100)}% business use."],
                  number_format=CURRENCY_FMT)
        r += 1

    # Personal
    if yd.get("personal_expenses", 0) > 0:
        write_row(ws, r, ["Personal (Non-Deductible)", "0%", "N/A", yd["personal_expenses"], 0, "Not deductible"])
        r += 1

    r += 1
    write_row(ws, r, ["TOTAL", "", "", total_spent, total_deductible, ""],
              font=Font(name='Arial', bold=True, size=11), fill=TOTAL_FILL, number_format=CURRENCY_FMT)

    r += 2

    # CCA section
    ws.cell(row=r, column=1, value="CAPITAL COST ALLOWANCE (CCA) -- Schedule 8").font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    r += 1
    cca_headers = ["Asset", "Class", "Rate", "Cost ($)", "Business %", "First Year CCA ($)"]
    apply_header(ws, r, cca_headers)
    r += 1

    for asset in yd["capital_assets"]:
        opening = yd["opening_ucc"].get(asset["class"], 0)
        business_cost = asset["cost"] * asset["business_pct"]
        cca = round(asset["rate"] * (opening + business_cost / 2), 2)
        write_row(ws, r, [
            asset["name"], f"Class {asset['class']}", f"{int(asset['rate']*100)}%",
            asset["cost"], f"{int(asset['business_pct']*100)}%", cca
        ], font=BLUE_INPUT, number_format=CURRENCY_FMT)
        r += 1

    # CC monthly breakdown
    if yd.get("cc_monthly"):
        r += 2
        ws.cell(row=r, column=1, value="MONTHLY CREDIT CARD BREAKDOWN").font = Font(name='Arial', bold=True, size=12, color='1F4E79')
        r += 1
        apply_header(ws, r, ["Month", "Charges ($)", "Credits ($)", "Net ($)", "", ""])
        r += 1
        for m in yd["cc_monthly"]:
            write_row(ws, r, [m["month"], m["charges"], m.get("credits", 0), m.get("net", 0), "", ""],
                      number_format=CURRENCY_FMT)
            r += 1

    auto_width(ws)
    ws.freeze_panes = 'A4'


def build_bank_statement(wb, data, year):
    """Sheet: Bank Statement for a specific year."""
    ws = wb.create_sheet(f"{year} Bank Statement")
    yd = data["years"][year]

    r = 1
    ws.cell(row=r, column=1, value=f"{data['company_name']} -- {year} Bank Account").font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    r += 1
    ws.cell(row=r, column=1, value=f"Opening: ${yd['bank_opening_balance']:,.2f}  |  Closing: ${yd['bank_closing_balance']:,.2f}").font = NAVY_SMALL
    r += 2

    headers = ["Month", "Date", "Description", "Withdrawal ($)", "Deposit ($)", "Category", "Tax Treatment"]
    apply_header(ws, r, headers)
    r += 1

    for txn in yd["bank_transactions"]:
        values = [
            txn.get("month", ""),
            txn.get("date", ""),
            txn.get("description", ""),
            txn.get("withdrawal"),
            txn.get("deposit"),
            txn.get("category", ""),
            txn.get("treatment", ""),
        ]
        is_flag = "FLAG" in txn.get("treatment", "") or "FLAG" in txn.get("category", "")
        fill = FLAG_FILL if is_flag else None
        write_row(ws, r, values, number_format=CURRENCY_FMT, fill=fill)
        r += 1

    auto_width(ws)
    ws.freeze_panes = 'A5'


def build_notes(wb, data):
    """Sheet: Notes & Action Items."""
    ws = wb.create_sheet("Notes & Action Items")

    r = 1
    ws.cell(row=r, column=1, value=f"Tax Filing Notes -- {data['company_name']}").font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    r += 2

    headers = ["Item", "Status", "Details"]
    apply_header(ws, r, headers)
    r += 1

    for item in data["action_items"]:
        write_row(ws, r, [item["item"], item["status"], item["details"]])
        r += 1

    if data.get("clarification_questions"):
        r += 2
        ws.cell(row=r, column=1, value="CLARIFICATION QUESTIONS").font = Font(name='Arial', bold=True, size=12, color='CC0000')
        r += 1
        for q in data["clarification_questions"]:
            ws.cell(row=r, column=1, value=q["question"]).font = BODY_FONT
            ws.cell(row=r, column=2, value=q.get("category", "")).font = BODY_FONT
            ws.cell(row=r, column=1).fill = FLAG_FILL
            r += 1

    auto_width(ws)


def generate_workbook(data):
    """Main entry point: generate the complete workbook."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets
    build_opening_position(wb, data)
    build_income_summary(wb, data)

    for year in data["fiscal_years"]:
        if year in data["years"]:
            build_tax_summary(wb, data, year)
            build_bank_statement(wb, data, year)

    build_notes(wb, data)

    # Save
    company_slug = data["company_name"].replace(" ", "_")
    years_str = "_".join(str(y) for y in data["fiscal_years"])
    filename = f"{company_slug}_Tax_{years_str}.xlsx"
    output_path = data["output_dir"].rstrip("/") + "/" + filename

    wb.save(output_path)
    print(f"Workbook saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    generate_workbook(DATA)
